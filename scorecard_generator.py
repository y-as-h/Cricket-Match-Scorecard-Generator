import openpyxl
import pandas as pd
import os
from datetime import datetime

start_time = datetime.now()

ind_innings = open("india_inns2.txt","r+")	#file for first innings
pak_innings = open("pak_inns1.txt","r+")	#file for second innings
teams = open("teams.txt","r+")			#file containing player name for both team

input_team = teams.readlines()

#####  list containing players of pakistan team
team_pakistan = input_team[0]
team_pakistan = team_pakistan[23:-1:].split(",")


#####  list containing players of india team
team_india = input_team[2]
team_india = team_india[20:-1:].split(",")

###removing blank lines
ind_inns=ind_innings.readlines()
for i in ind_inns:
    if i=='\n':
        ind_inns.remove(i)
      

pak_inns=pak_innings.readlines()
for i in pak_inns:
    if i=='\n':
        pak_inns.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active


pakistan_fow=0
pakistan_batsmen_out={}
pakistan_bowlers={}
pakistan_batsmen={}
pakistan_byes=0
total_pakistan_bowlers=0
india_fow=0
india_bowlers={}
india_batsmen={}

#### extracting data from pakistan innings
for line in pak_inns:
    x=line.index(".")
    pakistan_over=line[0:x+2]
    li=line[x+2::].split(",")
    current_ball=li[0].split("to") 

    if f"{current_ball[0].strip()}" not in india_bowlers.keys() :
        india_bowlers[f"{current_ball[0].strip()}"]=[1,0,0,0,0,0,0]   
    elif "wide" in li[1]:
        pass
    elif "bye" in li[1]:
        if "FOUR" in li[2]:
            pakistan_byes+=4
        elif "1" in li[2]:
            pakistan_byes+=1
        elif "2" in li[2]:
            pakistan_byes+=2
        elif "3" in li[2]:
            pakistan_byes+=3
        elif "4" in li[2]:
            pakistan_byes+=4
        elif "5" in li[2]:
            pakistan_byes+=5

    else:
        india_bowlers[f"{current_ball[0].strip()}"][0]+=1
    
    if f"{current_ball[1].strip()}" not in pakistan_batsmen.keys() and li[1]!="wide":
        pakistan_batsmen[f"{current_ball[1].strip()}"]=[0,1,0,0,0] 
    elif "wide" in li[1] :
        pass
    else:
        pakistan_batsmen[f"{current_ball[1].strip()}"][1]+=1
    

    if "out" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][3]+=1
        if "Bowled" in li[1].split("!!")[0]:
            pakistan_batsmen_out[f"{current_ball[1].strip()}"]=("b" + current_ball[0])
        elif "Caught" in li[1].split("!!")[0]:
            w=(li[1].split("!!")[0]).split("by")
            pakistan_batsmen_out[f"{current_ball[1].strip()}"]=("c" + w[1] +" b " + current_ball[0])
        elif "Lbw" in li[1].split("!!")[0]:
            pakistan_batsmen_out[f"{current_ball[1].strip()}"]=("lbw  b "+current_ball[0])

    

    if "no run" in li[1] or "out" in li[1] :
        india_bowlers[f"{current_ball[0].strip()}"][2]+=0
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=0
    elif "1 run" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=1
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=1
    elif "2 run" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=2
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=2
    elif "3 run" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=3
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=3
    elif "4 run" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=4
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=4
    elif "FOUR" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=4
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=4
        pakistan_batsmen[f"{current_ball[1].strip()}"][2]+=1
    elif "SIX" in li[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=6
        pakistan_batsmen[f"{current_ball[1].strip()}"][0]+=6
        pakistan_batsmen[f"{current_ball[1].strip()}"][3]+=1
    elif "wide" in li[1]:
        if "wides" in li[1]:
            india_bowlers[f"{current_ball[0].strip()}"][2]+=int(li[1][1])
            india_bowlers[f"{current_ball[0].strip()}"][5]+=int(li[1][1])
        else:
            india_bowlers[f"{current_ball[0].strip()}"][2]+=1
            india_bowlers[f"{current_ball[0].strip()}"][5]+=1

for score_card in pakistan_batsmen.values():
    score_card[-1]=round((score_card[0]/score_card[1])*100 , 2)



total_india_bowlers=0
india_byes=0
india_batsmen_out={}

### extracting data from india innings
for line in ind_inns:
    x=line.index(".")
    india_over=line[0:x+2]

    li=line[x+2::].split(",")

    current_ball=li[0].split("to") #0 2
    if f"{current_ball[0].strip()}" not in pakistan_bowlers.keys() :
        pakistan_bowlers[f"{current_ball[0].strip()}"]=[1,0,0,0,0,0,0]   
    elif "wide" in li[1]:
        pass
    elif "bye" in li[1]:
        if "FOUR" in li[2]:
            india_byes+=4
        elif "1" in li[2]:
            india_byes+=1
        elif "2" in li[2]:
            india_byes+=2
        elif "3" in li[2]:
            india_byes+=3
        elif "4" in li[2]:
            india_byes+=4
        elif "5" in li[2]:
            india_byes+=5
    else:
        pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
    
    if f"{current_ball[1].strip()}" not in india_batsmen.keys() and li[1]!="wide":
        india_batsmen[f"{current_ball[1].strip()}"]=[0,1,0,0,0] 
    elif "wide" in li[1] :
        pass
    else:
        india_batsmen[f"{current_ball[1].strip()}"][1]+=1
    

    if "out" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][3]+=1
        
        if "Bowled" in li[1].split("!!")[0]:
            india_batsmen_out[f"{current_ball[1].strip()}"]=("b" + current_ball[0])
        elif "Caught" in li[1].split("!!")[0]:
            w=(li[1].split("!!")[0]).split("by")
            india_batsmen_out[f"{current_ball[1].strip()}"]=("c" + w[1] +" b " + current_ball[0])
        elif "Lbw" in li[1].split("!!")[0]:
            india_batsmen_out[f"{current_ball[1].strip()}"]=("lbw  b "+current_ball[0])

    
    
    if "no run" in li[1] or "out" in li[1] :
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=0
        india_batsmen[f"{current_ball[1].strip()}"][0]+=0
    elif "1 run" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=1
        india_batsmen[f"{current_ball[1].strip()}"][0]+=1
    elif "2 run" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=2
        india_batsmen[f"{current_ball[1].strip()}"][0]+=2
    elif "3 run" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=3
        india_batsmen[f"{current_ball[1].strip()}"][0]+=3
    elif "4 run" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=4
        india_batsmen[f"{current_ball[1].strip()}"][0]+=4
    elif "FOUR" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=4
        india_batsmen[f"{current_ball[1].strip()}"][0]+=4
        india_batsmen[f"{current_ball[1].strip()}"][2]+=1
    elif "SIX" in li[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=6
        india_batsmen[f"{current_ball[1].strip()}"][0]+=6
        india_batsmen[f"{current_ball[1].strip()}"][3]+=1
    elif "wide" in li[1]:
        if "wides" in li[1]:
            pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=int(li[1][1])
            pakistan_bowlers[f"{current_ball[0].strip()}"][5]+=int(li[1][1])
        else:
            pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=1
            pakistan_bowlers[f"{current_ball[0].strip()}"][5]+=1


for score_card in india_batsmen.values():
    score_card[-1]=round((score_card[0]/score_card[1])*100 , 2)

for score_card in pakistan_batsmen.values():
    score_card[-1]=round((score_card[0]/score_card[1])*100 , 2)

for score_card in india_bowlers.values():
    if score_card[0]%6==0:
        score_card[0] = score_card[0]//6
    else:
        score_card[0] = (score_card[0]//6) + (score_card[0]%6)/10

for score_card in pakistan_bowlers.values():
    if score_card[0]%6==0:
        score_card[0] = score_card[0]//6
    else:
        score_card[0] = (score_card[0]//6) + (score_card[0]%6)/10

for score_card in india_bowlers.values(): 
    x=str(score_card[0])
    if "." in x:
        balls_bowled = int(x[0])*6 + int(x[2])
        score_card[-1]=round((score_card[2]/balls_bowled)*6,1)
    else:
        score_card[-1] = round((score_card[2]/score_card[0]) ,1) 


for score_card in pakistan_bowlers.values():
    x=str(score_card[0])
    if "." in x:
        balls_bowled = int(x[0])*6 + int(x[2])
        score_card[-1]=round((score_card[2]/balls_bowled)*6,1)
    else:
        score_card[-1] = round((score_card[2]/score_card[0]) ,1)



pakistan_batsmen_name=[]
for key in pakistan_batsmen.keys():
    pakistan_batsmen_name.append(key)

##### filling data in excel file
#### pakistan battting
for i in range(len(pakistan_batsmen)):
    sheet.cell(5+i,1).value = pakistan_batsmen_name[i]
    sheet.cell(5+i,5).value = pakistan_batsmen[pakistan_batsmen_name[i]][0]
    sheet.cell(5+i,6).value = pakistan_batsmen[pakistan_batsmen_name[i]][1]
    sheet.cell(5+i,7).value = pakistan_batsmen[pakistan_batsmen_name[i]][2]
    sheet.cell(5+i,8).value = pakistan_batsmen[pakistan_batsmen_name[i]][3]
    sheet.cell(5+i,9).value = pakistan_batsmen[pakistan_batsmen_name[i]][4]
    if pakistan_batsmen_name[i] not in pakistan_batsmen_out:
        sheet.cell(5+i,3).value = "not out"
    else:
        sheet.cell(5+i,3).value=pakistan_batsmen_out[pakistan_batsmen_name[i]]

sheet.cell(3,1).value = "BATTER"
sheet["E3"] = "RUNS"
sheet["F3"] = "BALLS"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "



sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER"
sheet["D18"] = "MAIDEN"
sheet["E18"] = "RUNS"
sheet["F18"] = "WICKET"
sheet["G18"] = "NO-BALL"
sheet["H18"] = "WIDE"
sheet["I18"] = "ECONOMY"

#### pakistan bowling
pakistan_bowlers_name=[]
for key in pakistan_bowlers.keys():
    pakistan_bowlers_name.append(key)

for i in range(len(pakistan_bowlers)):
    sheet.cell(42+i,1).value = pakistan_bowlers_name[i]
    sheet.cell(42+i,3).value = pakistan_bowlers[pakistan_bowlers_name[i]][0]
    sheet.cell(42+i,4).value = pakistan_bowlers[pakistan_bowlers_name[i]][1]
    sheet.cell(42+i,5).value = pakistan_bowlers[pakistan_bowlers_name[i]][2]
    sheet.cell(42+i,6).value = pakistan_bowlers[pakistan_bowlers_name[i]][3]
    sheet.cell(42+i,7).value = pakistan_bowlers[pakistan_bowlers_name[i]][4]
    sheet.cell(42+i,8).value = pakistan_bowlers[pakistan_bowlers_name[i]][5]
    sheet.cell(42+i,9).value = pakistan_bowlers[pakistan_bowlers_name[i]][6]
    total_pakistan_bowlers+=pakistan_bowlers[pakistan_bowlers_name[i]][2]
    india_fow+=pakistan_bowlers[pakistan_bowlers_name[i]][3]


sheet.cell(11+len(pakistan_batsmen)+len(pakistan_bowlers),1).value = "# INDIA"
sheet.cell(11+len(pakistan_batsmen)+len(pakistan_bowlers),2).value = " INNINGS"

india_batsmen_name=[]
for key in india_batsmen.keys():
    india_batsmen_name.append(key)


#### india batting
for i in range(len(india_batsmen)):
    sheet.cell(31+i,1).value = india_batsmen_name[i]
    sheet.cell(31+i,5).value = india_batsmen[india_batsmen_name[i]][0]
    sheet.cell(31+i,6).value = india_batsmen[india_batsmen_name[i]][1]
    sheet.cell(31+i,7).value = india_batsmen[india_batsmen_name[i]][2]
    sheet.cell(31+i,8).value = india_batsmen[india_batsmen_name[i]][3]
    sheet.cell(31+i,9).value = india_batsmen[india_batsmen_name[i]][4]

    if india_batsmen_name[i] not in india_batsmen_out:
        sheet.cell(31+i,3).value = "not out"
    else:
        sheet.cell(31+i,3).value=india_batsmen_out[india_batsmen_name[i]]

sheet["A29"] = "BATTER"
sheet["E29"] = "RUNS"
sheet["F29"] = "BALLS"
sheet["G29"] = " 4s "
sheet["H29"] = " 6s "
sheet["I29"] = "  SR  "



sheet["A40"] = "BOWLER"
sheet["C40"] = "OVER"
sheet["D40"] = "MAIDEN"
sheet["E40"] = "RUNS"
sheet["F40"] = "WICKET"
sheet["G40"] = "NO-BALL"
sheet["H40"] = "WIDE"
sheet["I40"] = "ECONOMY"

#### india bowling
india_bowlers_name=[]
for key in india_bowlers.keys():
    india_bowlers_name.append(key)

for i in range(len(india_bowlers)):

    sheet.cell(20+i,1).value = india_bowlers_name[i]
    sheet.cell(20+i,3).value = india_bowlers[india_bowlers_name[i]][0]
    sheet.cell(20+i,4).value = india_bowlers[india_bowlers_name[i]][1]
    sheet.cell(20+i,5).value = india_bowlers[india_bowlers_name[i]][2]
    sheet.cell(20+i,6).value = india_bowlers[india_bowlers_name[i]][3]
    sheet.cell(20+i,7).value = india_bowlers[india_bowlers_name[i]][4]
    sheet.cell(20+i,8).value = india_bowlers[india_bowlers_name[i]][5]
    sheet.cell(20+i,9).value = india_bowlers[india_bowlers_name[i]][6]
    total_india_bowlers+=india_bowlers[india_bowlers_name[i]][2]
    pakistan_fow+=india_bowlers[india_bowlers_name[i]][3]

india_score=total_india_bowlers+pakistan_byes
pakistan_score = total_pakistan_bowlers+india_byes

sheet["H27"] = " "+str(india_score) +" - " + str(india_fow)
sheet["I27"] = str(india_over)
Score=" "+str(pakistan_score) +" - " + str(pakistan_fow)
Over = str(pakistan_over)

####saving workbook
wb.save("Scoreboard.xlsx")

df = pd.read_excel('Scoreboard.xlsx')

df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" "," "," "," "," ",Score,Over], axis='columns')

### converting excel to csv
df.to_csv('Scorecard.csv',index=False)


#### using try except block for deleting scoreboard.xlsx
try:
    os.path.exists("Scoreboard.xlsx") 
    os.remove("Scoreboard.xlsx") 
except:
    print("Extra created file does not exist")

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
